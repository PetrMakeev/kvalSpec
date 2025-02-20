from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font

def export_xls(data, file_xls):
    # Исходные данные

    # Создаем Excel-файл
    wb = Workbook()
    ws = wb.active

    # Устанавливаем заголовок листа
    ws.title = "Специалисты"

    # Добавляем заголовки колонок
    headers = ["№ п/п", 
            "Фамилия И.О.", 
            "Подразделение", 
            "Должность", 
            "Звание", 
            "Квалификационное звание", 
            "Приказ о присвоении квалификационного звания", 
            "Дата следующего испытания", 
            "Примечание"]
    ws.append(headers)

    # Стили для обводки ячеек
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Форматирование заголовков
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Обработка заголовков
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Обработка заголовков
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index)
        cell.value = header
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Устанавливаем высоту первой строки
    ws.row_dimensions[1].height = 65  # Устанавливаем высоту в 65 пунктов        

    # Добавляем данные и обводим ячейки
    for row_index, row_data in enumerate(data, start=2):  # Начинаем с 2, т.к. 1 - заголовок
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_index, column=col_index)
            cell.value = value
            cell.border = thin_border
            
            # Выравнивание данных по центру для колонок F и H
            if col_index in [6, 8]:  # F = 6
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Настройка ширины колонок F и G
    ws.column_dimensions["A"].width = 5  
    ws.column_dimensions["B"].width = 35  
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 30  
    ws.column_dimensions["F"].width = 15
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 15          
    ws.column_dimensions["I"].width = 15  
    

    # Сохраняем файл

    wb.save(file_xls)


